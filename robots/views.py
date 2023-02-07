from openpyxl.styles import Alignment
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from rest_framework.generics import ListAPIView
from django.http import HttpResponse
from django.db.models import Count
from datetime import datetime, timedelta

from robots.models import Robot


class ReportViewSet(ListAPIView):

    def get(self, request, format=None):
        report_period_from = datetime.now() - timedelta(weeks=1)
        report_period_from.strftime("%Y-%m-%d %H:%M:%S")
        wb = Workbook()
        records = Robot.objects.values('model').distinct()
        for row in records:
            rec = Robot.objects \
                .filter(model=row['model'], created__gt=report_period_from)\
                .values('model', 'version') \
                .annotate(total=Count('version'))
            ws = wb.create_sheet(row['model'])
            ws['A1'] = 'Модель'
            ws['B1'] = 'Версия'
            ws['C1'] = 'Количество за неделю'
            ws.column_dimensions['C'].width = 21
            for data in rec:
                ws.append(list(data.values()))
            for item in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=3):
                for cell in item:
                    cell.alignment = Alignment(horizontal='center')
        std = wb.get_sheet_by_name('Sheet')
        wb.remove_sheet(std)
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename=Production_report.xlsx'
        return response
